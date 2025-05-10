import openpyxl
from openpyxl.styles import PatternFill

def  getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columno).value

def writeData(file,sheetName,rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)